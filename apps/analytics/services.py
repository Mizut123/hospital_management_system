"""
Report Generation Service for Hospital Management System.

Provides PDF and Excel report generation using ReportLab and openpyxl.
"""
import io
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Count, Sum, Avg, F

# PDF Generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# Excel Generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ReportGenerator:
    """Generates various hospital reports in PDF and Excel formats."""

    def __init__(self):
        self.hospital_name = "Hospital Management System"
        self.generated_at = timezone.now()

    def generate_patient_report(self, format='pdf', date_from=None, date_to=None):
        """Generate patient statistics report."""
        from apps.accounts.models import User

        date_from = date_from or (timezone.now() - timedelta(days=30)).date()
        date_to = date_to or timezone.now().date()

        # Gather data
        patients = User.objects.filter(role=User.Role.PATIENT)
        total_patients = patients.count()
        new_patients = patients.filter(date_joined__date__gte=date_from, date_joined__date__lte=date_to).count()

        # Age distribution
        age_distribution = []
        today = timezone.now().date()
        for label, min_age, max_age in [('0-18', 0, 18), ('19-35', 19, 35), ('36-55', 36, 55), ('56-70', 56, 70), ('70+', 70, 200)]:
            count = patients.filter(
                date_of_birth__isnull=False
            ).extra(
                where=[f"CAST((julianday('now') - julianday(date_of_birth))/365 AS INTEGER) >= {min_age}",
                       f"CAST((julianday('now') - julianday(date_of_birth))/365 AS INTEGER) <= {max_age}"]
            ).count() if min_age != 70 else patients.filter(
                date_of_birth__isnull=False
            ).extra(
                where=[f"CAST((julianday('now') - julianday(date_of_birth))/365 AS INTEGER) >= {min_age}"]
            ).count()
            age_distribution.append((label, count))

        # Gender distribution
        gender_distribution = list(patients.values('gender').annotate(count=Count('id')))

        data = {
            'title': 'Patient Statistics Report',
            'date_range': f"{date_from} to {date_to}",
            'total_patients': total_patients,
            'new_patients': new_patients,
            'age_distribution': age_distribution,
            'gender_distribution': gender_distribution,
        }

        if format == 'pdf':
            return self._generate_patient_pdf(data)
        else:
            return self._generate_patient_excel(data)

    def generate_appointment_report(self, format='pdf', date_from=None, date_to=None):
        """Generate appointment statistics report."""
        from apps.appointments.models import Appointment
        from apps.accounts.models import Department

        date_from = date_from or (timezone.now() - timedelta(days=30)).date()
        date_to = date_to or timezone.now().date()

        appointments = Appointment.objects.filter(
            scheduled_date__gte=date_from,
            scheduled_date__lte=date_to
        )

        total = appointments.count()
        by_status = list(appointments.values('status').annotate(count=Count('id')))
        by_department = list(appointments.values('department__name').annotate(count=Count('id')).order_by('-count')[:10])

        # Daily trend
        daily_trend = []
        current = date_from
        while current <= date_to:
            count = appointments.filter(scheduled_date=current).count()
            daily_trend.append((current.strftime('%Y-%m-%d'), count))
            current += timedelta(days=1)

        data = {
            'title': 'Appointment Statistics Report',
            'date_range': f"{date_from} to {date_to}",
            'total_appointments': total,
            'by_status': by_status,
            'by_department': by_department,
            'daily_trend': daily_trend[-14:],  # Last 14 days
        }

        if format == 'pdf':
            return self._generate_appointment_pdf(data)
        else:
            return self._generate_appointment_excel(data)

    def generate_pharmacy_report(self, format='pdf', date_from=None, date_to=None):
        """Generate pharmacy/inventory report."""
        from apps.pharmacy.models import Medicine, MedicineStock, MedicineCategory, StockTransaction

        date_from = date_from or (timezone.now() - timedelta(days=30)).date()
        date_to = date_to or timezone.now().date()

        # Stock summary
        stocks = MedicineStock.objects.select_related('medicine')
        total_items = stocks.count()
        in_stock = stocks.filter(quantity__gt=F('reorder_level')).count()
        low_stock = stocks.filter(quantity__gt=0, quantity__lte=F('reorder_level')).count()
        out_of_stock = stocks.filter(quantity=0).count()

        # Top medicines by usage
        transactions = StockTransaction.objects.filter(
            transaction_type='out',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        )
        top_medicines = list(transactions.values(
            'stock__medicine__name'
        ).annotate(
            total_dispensed=Sum('quantity')
        ).order_by('-total_dispensed')[:10])

        # Low stock items
        low_stock_items = list(MedicineStock.objects.filter(
            quantity__lte=F('reorder_level')
        ).select_related('medicine').values(
            'medicine__name', 'quantity', 'reorder_level'
        )[:15])

        # Stock value estimate
        total_value = stocks.aggregate(
            value=Sum(F('quantity') * F('unit_price'))
        )['value'] or 0

        data = {
            'title': 'Pharmacy & Inventory Report',
            'date_range': f"{date_from} to {date_to}",
            'total_items': total_items,
            'in_stock': in_stock,
            'low_stock': low_stock,
            'out_of_stock': out_of_stock,
            'total_value': total_value,
            'top_medicines': top_medicines,
            'low_stock_items': low_stock_items,
        }

        if format == 'pdf':
            return self._generate_pharmacy_pdf(data)
        else:
            return self._generate_pharmacy_excel(data)

    def generate_doctor_activity_report(self, format='pdf', date_from=None, date_to=None):
        """Generate doctor activity and performance report."""
        from apps.appointments.models import Appointment
        from apps.accounts.models import User
        from apps.medical_records.models import MedicalRecord

        date_from = date_from or (timezone.now() - timedelta(days=30)).date()
        date_to = date_to or timezone.now().date()

        # Get all doctors with activity
        doctors = User.objects.filter(role=User.Role.DOCTOR, is_active=True)

        doctor_stats = []
        for doctor in doctors:
            appointments = Appointment.objects.filter(
                doctor=doctor,
                scheduled_date__gte=date_from,
                scheduled_date__lte=date_to
            )
            total = appointments.count()
            completed = appointments.filter(status='completed').count()
            cancelled = appointments.filter(status='cancelled').count()
            no_show = appointments.filter(status='no_show').count()

            # Medical records created
            records = MedicalRecord.objects.filter(
                doctor=doctor,
                visit_date__gte=date_from,
                visit_date__lte=date_to
            ).count()

            completion_rate = round((completed / total * 100), 1) if total > 0 else 0

            doctor_stats.append({
                'name': doctor.get_full_name() or doctor.email,
                'department': doctor.doctorprofile.department.name if hasattr(doctor, 'doctorprofile') and doctor.doctorprofile.department else 'Unassigned',
                'total_appointments': total,
                'completed': completed,
                'cancelled': cancelled,
                'no_show': no_show,
                'records_created': records,
                'completion_rate': completion_rate,
            })

        # Sort by total appointments
        doctor_stats.sort(key=lambda x: x['total_appointments'], reverse=True)

        # Summary stats
        total_doctors = len(doctor_stats)
        total_appointments = sum(d['total_appointments'] for d in doctor_stats)
        total_completed = sum(d['completed'] for d in doctor_stats)
        avg_completion_rate = round(sum(d['completion_rate'] for d in doctor_stats) / total_doctors, 1) if total_doctors > 0 else 0

        data = {
            'title': 'Doctor Activity Report',
            'date_range': f"{date_from} to {date_to}",
            'total_doctors': total_doctors,
            'total_appointments': total_appointments,
            'total_completed': total_completed,
            'avg_completion_rate': avg_completion_rate,
            'doctor_stats': doctor_stats[:20],  # Top 20
        }

        if format == 'pdf':
            return self._generate_doctor_activity_pdf(data)
        else:
            return self._generate_doctor_activity_excel(data)

    # PDF Generation Methods
    def _generate_patient_pdf(self, data):
        if not HAS_REPORTLAB:
            raise ImportError("ReportLab is required for PDF generation. Install with: pip install reportlab")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=20)
        elements.append(Paragraph(data['title'], title_style))
        elements.append(Paragraph(f"Period: {data['date_range']}", styles['Normal']))
        elements.append(Paragraph(f"Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary
        summary_data = [
            ['Metric', 'Value'],
            ['Total Patients', str(data['total_patients'])],
            ['New Patients (Period)', str(data['new_patients'])],
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Age Distribution
        elements.append(Paragraph("Age Distribution", styles['Heading2']))
        age_data = [['Age Group', 'Count']] + [[age, str(count)] for age, count in data['age_distribution']]
        age_table = Table(age_data, colWidths=[2*inch, 2*inch])
        age_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(age_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _generate_appointment_pdf(self, data):
        if not HAS_REPORTLAB:
            raise ImportError("ReportLab is required for PDF generation")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=20)
        elements.append(Paragraph(data['title'], title_style))
        elements.append(Paragraph(f"Period: {data['date_range']}", styles['Normal']))
        elements.append(Paragraph(f"Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary
        elements.append(Paragraph(f"Total Appointments: {data['total_appointments']}", styles['Heading2']))
        elements.append(Spacer(1, 10))

        # By Status
        elements.append(Paragraph("Appointments by Status", styles['Heading3']))
        status_data = [['Status', 'Count']] + [[s['status'].title(), str(s['count'])] for s in data['by_status']]
        status_table = Table(status_data, colWidths=[2.5*inch, 1.5*inch])
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 20))

        # By Department
        elements.append(Paragraph("Appointments by Department", styles['Heading3']))
        dept_data = [['Department', 'Count']] + [[d['department__name'] or 'Unassigned', str(d['count'])] for d in data['by_department']]
        dept_table = Table(dept_data, colWidths=[3*inch, 1.5*inch])
        dept_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(dept_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _generate_pharmacy_pdf(self, data):
        if not HAS_REPORTLAB:
            raise ImportError("ReportLab is required for PDF generation")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=20)
        elements.append(Paragraph(data['title'], title_style))
        elements.append(Paragraph(f"Period: {data['date_range']}", styles['Normal']))
        elements.append(Paragraph(f"Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Stock Summary
        summary_data = [
            ['Metric', 'Value'],
            ['Total Stock Items', str(data['total_items'])],
            ['In Stock', str(data['in_stock'])],
            ['Low Stock', str(data['low_stock'])],
            ['Out of Stock', str(data['out_of_stock'])],
            ['Estimated Value', f"${data['total_value']:,.2f}"],
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#FEF3C7')),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#FEE2E2')),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Low Stock Items
        if data['low_stock_items']:
            elements.append(Paragraph("Low Stock Alert", styles['Heading2']))
            low_data = [['Medicine', 'Current', 'Reorder Level']]
            for item in data['low_stock_items']:
                low_data.append([
                    item['medicine__name'],
                    str(item['quantity']),
                    str(item['reorder_level'])
                ])
            low_table = Table(low_data, colWidths=[2.5*inch, 1*inch, 1.5*inch])
            low_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(low_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _generate_doctor_activity_pdf(self, data):
        if not HAS_REPORTLAB:
            raise ImportError("ReportLab is required for PDF generation")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=20)
        elements.append(Paragraph(data['title'], title_style))
        elements.append(Paragraph(f"Period: {data['date_range']}", styles['Normal']))
        elements.append(Paragraph(f"Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary Stats
        summary_data = [
            ['Metric', 'Value'],
            ['Active Doctors', str(data['total_doctors'])],
            ['Total Appointments', str(data['total_appointments'])],
            ['Completed', str(data['total_completed'])],
            ['Avg Completion Rate', f"{data['avg_completion_rate']}%"],
        ]
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C3AED')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Doctor Activity Table
        elements.append(Paragraph("Doctor Performance", styles['Heading2']))
        doctor_data = [['Doctor', 'Dept', 'Total', 'Done', 'Rate']]
        for doc_stat in data['doctor_stats']:
            doctor_data.append([
                doc_stat['name'][:20],
                doc_stat['department'][:15],
                str(doc_stat['total_appointments']),
                str(doc_stat['completed']),
                f"{doc_stat['completion_rate']}%"
            ])

        if len(doctor_data) > 1:
            doctor_table = Table(doctor_data, colWidths=[1.8*inch, 1.3*inch, 0.7*inch, 0.7*inch, 0.7*inch])
            doctor_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(doctor_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    # Excel Generation Methods
    def _generate_patient_excel(self, data):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Patient Report"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = data['title']
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Period: {data['date_range']}"
        ws['A3'] = f"Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M')}"

        # Summary
        ws['A5'] = 'Metric'
        ws['B5'] = 'Value'
        ws['A5'].font = header_font
        ws['B5'].font = header_font
        ws['A5'].fill = header_fill
        ws['B5'].fill = header_fill

        ws['A6'] = 'Total Patients'
        ws['B6'] = data['total_patients']
        ws['A7'] = 'New Patients (Period)'
        ws['B7'] = data['new_patients']

        # Age Distribution
        ws['A10'] = 'Age Distribution'
        ws['A10'].font = Font(bold=True, size=12)

        ws['A11'] = 'Age Group'
        ws['B11'] = 'Count'
        ws['A11'].font = header_font
        ws['B11'].font = header_font
        ws['A11'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        ws['B11'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

        for i, (age_group, count) in enumerate(data['age_distribution'], start=12):
            ws[f'A{i}'] = age_group
            ws[f'B{i}'] = count

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _generate_appointment_excel(self, data):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel generation")

        wb = Workbook()
        ws = wb.active
        ws.title = "Appointments Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")

        ws['A1'] = data['title']
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Period: {data['date_range']}"
        ws['A3'] = f"Total Appointments: {data['total_appointments']}"

        # By Status
        ws['A5'] = 'By Status'
        ws['A5'].font = Font(bold=True, size=12)

        ws['A6'] = 'Status'
        ws['B6'] = 'Count'
        ws['A6'].font = header_font
        ws['B6'].font = header_font
        ws['A6'].fill = header_fill
        ws['B6'].fill = header_fill

        for i, item in enumerate(data['by_status'], start=7):
            ws[f'A{i}'] = item['status'].title()
            ws[f'B{i}'] = item['count']

        # By Department
        start_row = 7 + len(data['by_status']) + 2
        ws[f'A{start_row}'] = 'By Department'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)

        ws[f'A{start_row+1}'] = 'Department'
        ws[f'B{start_row+1}'] = 'Count'
        ws[f'A{start_row+1}'].font = header_font
        ws[f'B{start_row+1}'].font = header_font
        ws[f'A{start_row+1}'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        ws[f'B{start_row+1}'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")

        for i, item in enumerate(data['by_department'], start=start_row+2):
            ws[f'A{i}'] = item['department__name'] or 'Unassigned'
            ws[f'B{i}'] = item['count']

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _generate_pharmacy_excel(self, data):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel generation")

        wb = Workbook()
        ws = wb.active
        ws.title = "Pharmacy Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")

        ws['A1'] = data['title']
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Period: {data['date_range']}"

        # Summary
        ws['A4'] = 'Stock Summary'
        ws['A4'].font = Font(bold=True, size=12)

        summary_items = [
            ('Total Items', data['total_items']),
            ('In Stock', data['in_stock']),
            ('Low Stock', data['low_stock']),
            ('Out of Stock', data['out_of_stock']),
            ('Estimated Value', f"${data['total_value']:,.2f}"),
        ]

        for i, (label, value) in enumerate(summary_items, start=5):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value

        # Low Stock Items
        start_row = 5 + len(summary_items) + 2
        ws[f'A{start_row}'] = 'Low Stock Items'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)

        ws[f'A{start_row+1}'] = 'Medicine'
        ws[f'B{start_row+1}'] = 'Current'
        ws[f'C{start_row+1}'] = 'Reorder Level'
        for col in ['A', 'B', 'C']:
            ws[f'{col}{start_row+1}'].font = header_font
            ws[f'{col}{start_row+1}'].fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

        for i, item in enumerate(data['low_stock_items'], start=start_row+2):
            ws[f'A{i}'] = item['medicine__name']
            ws[f'B{i}'] = item['quantity']
            ws[f'C{i}'] = item['reorder_level']

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _generate_doctor_activity_excel(self, data):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel generation")

        wb = Workbook()
        ws = wb.active
        ws.title = "Doctor Activity"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")

        ws['A1'] = data['title']
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Period: {data['date_range']}"

        # Summary
        ws['A4'] = 'Summary'
        ws['A4'].font = Font(bold=True, size=12)

        summary_items = [
            ('Active Doctors', data['total_doctors']),
            ('Total Appointments', data['total_appointments']),
            ('Completed', data['total_completed']),
            ('Avg Completion Rate', f"{data['avg_completion_rate']}%"),
        ]

        for i, (label, value) in enumerate(summary_items, start=5):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value

        # Doctor Stats Table
        start_row = 5 + len(summary_items) + 2
        ws[f'A{start_row}'] = 'Doctor Performance'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)

        headers = ['Doctor', 'Department', 'Total', 'Completed', 'Cancelled', 'No Show', 'Records', 'Rate %']
        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f'{col}{start_row+1}'] = header
            ws[f'{col}{start_row+1}'].font = header_font
            ws[f'{col}{start_row+1}'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

        for i, doc_stat in enumerate(data['doctor_stats'], start=start_row+2):
            ws[f'A{i}'] = doc_stat['name']
            ws[f'B{i}'] = doc_stat['department']
            ws[f'C{i}'] = doc_stat['total_appointments']
            ws[f'D{i}'] = doc_stat['completed']
            ws[f'E{i}'] = doc_stat['cancelled']
            ws[f'F{i}'] = doc_stat['no_show']
            ws[f'G{i}'] = doc_stat['records_created']
            ws[f'H{i}'] = doc_stat['completion_rate']

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
